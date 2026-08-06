"""Query result exporter module.

Supports exporting query results to CSV, JSON, and Excel formats.
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Custom exception for export errors."""
    pass


class ExporterResult:
    """Result of an export operation."""

    def __init__(self, file_path: str, format: str, row_count: int, message: str = ""):
        self.file_path = file_path
        self.format = format
        self.row_count = row_count
        self.message = message

    def to_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "format": self.format,
            "row_count": self.row_count,
            "message": self.message,
        }


class QueryExporter:
    """Export query results to various formats.

    Supported formats: CSV, JSON, Excel (.xlsx)
    """

    def __init__(self, export_dir: Optional[str] = None):
        """Initialize exporter.

        Args:
            export_dir: Directory for exported files. Defaults to ~/.db_query/exports
        """
        if export_dir:
            self.export_dir = Path(export_dir)
        else:
            self.export_dir = Path.home() / ".db_query" / "exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        columns: List[str],
        rows: List[Dict[str, Any]],
        format: str,
        filename: Optional[str] = None,
    ) -> ExporterResult:
        """Export query results to specified format.

        Args:
            columns: List of column names
            rows: List of row dictionaries
            format: Export format ('csv', 'json', 'excel')
            filename: Optional filename (without extension)

        Returns:
            ExporterResult with file path and metadata

        Raises:
            ExportError: If export fails
            ValueError: If format is unsupported
        """
        format = format.lower().strip()
        if format not in ("csv", "json", "excel", "xlsx"):
            raise ValueError(f"Unsupported export format: {format}")

        if not rows:
            raise ExportError("No data to export. The query returned no results.")

        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"query_result_{timestamp}"

        filename = self._sanitize_filename(filename)

        try:
            if format == "csv":
                return self._export_csv(columns, rows, filename)
            elif format == "json":
                return self._export_json(columns, rows, filename)
            elif format in ("excel", "xlsx"):
                return self._export_excel(columns, rows, filename)
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            raise ExportError(f"Failed to export data: {str(e)}") from e

    def _export_csv(
        self, columns: List[str], rows: List[Dict[str, Any]], filename: str
    ) -> ExporterResult:
        """Export to CSV format."""
        file_path = self.export_dir / f"{filename}.csv"

        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"CSV exported to: {file_path}")
        return ExporterResult(
            file_path=str(file_path),
            format="csv",
            row_count=len(rows),
            message=f"Successfully exported {len(rows)} rows to CSV.",
        )

    def _export_json(
        self, columns: List[str], rows: List[Dict[str, Any]], filename: str
    ) -> ExporterResult:
        """Export to JSON format."""
        file_path = self.export_dir / f"{filename}.json"

        export_data = {
            "columns": columns,
            "rowCount": len(rows),
            "exportedAt": datetime.now().isoformat(),
            "data": rows,
        }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)

        logger.info(f"JSON exported to: {file_path}")
        return ExporterResult(
            file_path=str(file_path),
            format="json",
            row_count=len(rows),
            message=f"Successfully exported {len(rows)} rows to JSON.",
        )

    def _export_excel(
        self, columns: List[str], rows: List[Dict[str, Any]], filename: str
    ) -> ExporterResult:
        """Export to Excel format with styles.

        Excel requirements:
        - Header row (Row 1): Dark yellow background, Microsoft YaHei font, bold, black borders
        - Data rows: Microsoft YaHei font, not bold, black borders
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise ExportError(
                "Excel export requires 'openpyxl' library. "
                "Install it with: pip install openpyxl"
            )

        from app.services.exporter.styles import (
            HEADER_FONT,
            HEADER_FILL,
            DATA_FONT,
            BLACK_BORDER,
            CELL_ALIGNMENT,
        )

        file_path = self.export_dir / f"{filename}.xlsx"

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Results"

        # Write header row (Row 1)
        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BLACK_BORDER
            cell.alignment = CELL_ALIGNMENT

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, column_name in enumerate(columns, start=1):
                value = row_data.get(column_name, "")
                # Convert non-serializable types to string
                if isinstance(value, (list, dict, tuple, set)):
                    value = str(value)
                elif value is None:
                    value = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = BLACK_BORDER
                cell.alignment = CELL_ALIGNMENT

        # Auto-adjust column widths
        for col_idx, column_name in enumerate(columns, start=1):
            max_length = len(str(column_name))
            for row_data in rows:
                cell_value = row_data.get(column_name, "")
                cell_length = len(str(cell_value)) if cell_value else 0
                max_length = max(max_length, cell_length)
            # Set column width with some padding, cap at 50
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save file
        wb.save(file_path)

        logger.info(f"Excel exported to: {file_path}")
        return ExporterResult(
            file_path=str(file_path),
            format="excel",
            row_count=len(rows),
            message=f"Successfully exported {len(rows)} rows to Excel with styles.",
        )

    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        """Sanitize filename by removing invalid characters."""
        # Remove invalid characters for Windows/macOS/Linux
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, "")
        # Ensure filename doesn't start with dot or space
        filename = filename.strip(". ")
        # Truncate to reasonable length
        if len(filename) > 100:
            filename = filename[:100]
        # Ensure filename is not empty
        if not filename:
            filename = "query_result"
        return filename

    def get_export_dir(self) -> str:
        """Get the export directory path."""
        return str(self.export_dir)

    @staticmethod
    def recommend_format(columns: List[str], rows: List[Dict[str, Any]]) -> str:
        """Recommend export format based on data characteristics.

        Args:
            columns: List of column names
            rows: List of row data

        Returns:
            Recommended format: 'excel', 'json', or 'csv'
        """
        if not rows:
            return "csv"

        # Check for Chinese characters or complex data
        has_chinese = False
        has_complex_data = False
        total_complex_cells = 0
        total_cells = 0

        for col in columns:
            if any("\u4e00" <= char <= "\u9fff" for char in col):
                has_chinese = True
                break

        for row in rows[:50]:  # Sample first 50 rows
            for col in columns:
                value = row.get(col)
                if value is None:
                    continue
                total_cells += 1
                if any("\u4e00" <= char <= "\u9fff" for char in str(value)):
                    has_chinese = True
                if isinstance(value, (dict, list, tuple, set)):
                    has_complex_data = True
                    total_complex_cells += 1

        # High column count or Chinese text -> Excel
        if len(columns) > 10 or has_chinese:
            return "excel"

        # Complex nested data -> JSON
        if has_complex_data and total_complex_cells > 0:
            return "json"

        # Default for pure numeric/simple data -> CSV
        return "csv"
